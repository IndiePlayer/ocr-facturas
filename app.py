import os
import re
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

app = Flask(__name__)
app.secret_key = "clave_secreta_ocr_facturas"

UPLOAD_FOLDER = "uploads"
EXCEL_FILE = "facturas.xlsx"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "bmp", "tiff", "tif", "webp"}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

# ──────────────────────────────────────────────
# Utilidades
# ──────────────────────────────────────────────

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def preprocesar_imagen(img):
    img = img.convert("L")
    ancho, alto = img.size
    img = img.resize((ancho * 2, alto * 2), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(2.5)
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    img = img.point(lambda x: 0 if x < 140 else 255)
    return img


def extraer_texto_ocr(filepath):
    try:
        img = Image.open(filepath)
        img = preprocesar_imagen(img)
        config = "--oem 3 --psm 6 -l spa+eng"
        texto = pytesseract.image_to_string(img, config=config)
        return texto
    except Exception as e:
        return f"ERROR_OCR: {str(e)}"


def extraer_datos(texto):
    resultado = {"fecha": None, "codigo_aprobado": None, "total": None}

    # ── CÓDIGO DE APROBADO ──────────────────────────────────────────────────
    patrones_codigo = [
        r"APRO\.?\s*[:#]?\s*(\d{4,10})",
        r"(?:c[oó]d(?:igo)?\.?\s*(?:de\s*)?aprobaci[oó]n|aprobado|aprobaci[oó]n|auth(?:orization)?|autorizaci[oó]n)\s*[:#\-]?\s*([A-Z0-9\-]{4,20})",
        r"(?:no\.?\s*aprobaci[oó]n|n[uú]mero\s*aprobaci[oó]n)\s*[:#\-]?\s*([A-Z0-9\-]{4,20})",
        r"\bAPROB(?:ADO)?\b[\s:]*([A-Z0-9\-]{4,20})",
    ]
    for patron in patrones_codigo:
        match = re.search(patron, texto, re.IGNORECASE)
        if match:
            resultado["codigo_aprobado"] = match.group(1).strip()
            break

    # ── TOTAL ────────────────────────────────────────────────────────────────
    patrones_total = [
        r"(?:TOTAL|'OTAL|OTAL|T0TAL)\s*\$?\s*([\d\.\,]+)",
        r"(?:COMPRA\s*NETA)\s*\$?\s*([\d\s\.\,]+)",
        r"(?:total\s*(?:a\s*pagar|factura|general|importe)?|gran\s*total|importe\s*total|monto\s*total)\s*[:\$]?\s*\$?\s*([\d\.,]+)",
        r"(?:valor\s*total|total\s*bs\.?|total\s*cop)\s*[:\$]?\s*([\d\.,]+)",
    ]
    for patron in patrones_total:
        match = re.search(patron, texto, re.IGNORECASE)
        if match:
            total_limpio = re.sub(r"[\[\]\(\)]+$", "", match.group(1)).strip()
            total_limpio = re.sub(r"\s+", "", total_limpio)
            resultado["total"] = total_limpio
            break

    # ── FECHA ────────────────────────────────────────────────────────────────
    texto_fecha = texto
    texto_fecha = re.sub(r'\b200(\d)[/\-](\d{2})[/\-](\d{2})', r'202\1/\2/\3', texto_fecha)

    patrones_fecha = [
        r"\b(202[0-9][/\-]\d{1,2}[/\-]\d{1,2})\b",
        r"\b(\d{1,2}[/\-]\d{1,2}[/\-]202[0-9])\b",
        r"\b(\d{1,2}\s+de\s+(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:de\s+)?20\d{2})\b",
        r"\b(\d{1,2}\s+(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+20\d{2})\b",
        r"\b((?:january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{1,2},?\s+20\d{2})\b",
        r"\b(20\d{2}[/\-]\d{1,2}[/\-]\d{1,2})\b",
    ]
    for patron in patrones_fecha:
        match = re.search(patron, texto_fecha, re.IGNORECASE)
        if match:
            resultado["fecha"] = match.group(1).strip()
            break

    return resultado


# ──────────────────────────────────────────────
# Excel
# ──────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
DATA_FONT         = Font(name="Arial", size=10)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


def crear_excel_si_no_existe():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"
        encabezados = ["Fecha", "Código de Aprobado", "Valor Total"]
        for col, titulo in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border    = THIN_BORDER
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.row_dimensions[1].height = 30
        wb.save(EXCEL_FILE)


def guardar_en_excel(fecha, codigo_aprobado, total):
    crear_excel_si_no_existe()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Facturas"]
    nueva_fila = ws.max_row + 1

    valores      = [fecha or "No detectado", codigo_aprobado or "No detectado", total or "No detectado"]
    alineaciones = [DATA_ALIGN_CENTER, DATA_ALIGN_CENTER, DATA_ALIGN_RIGHT]

    for col, (val, aln) in enumerate(zip(valores, alineaciones), start=1):
        cell           = ws.cell(row=nueva_fila, column=col, value=val)
        cell.font      = DATA_FONT
        cell.alignment = aln
        cell.border    = THIN_BORDER
        if nueva_fila % 2 == 0:
            cell.fill = PatternFill("solid", start_color="EBF3FB")

    wb.save(EXCEL_FILE)


# ──────────────────────────────────────────────
# Rutas Flask
# ──────────────────────────────────────────────

@app.route("/", methods=["GET"])
def index():
    registros = obtener_registros()
    return render_template("index.html", registros=registros)


@app.route("/upload", methods=["POST"])
def upload():
    archivos = request.files.getlist("archivos")
    if not archivos or all(f.filename == "" for f in archivos):
        flash("⚠️ No seleccionaste ningún archivo.", "warning")
        return redirect(url_for("index"))

    resultados = []
    for archivo in archivos:
        if archivo and allowed_file(archivo.filename):
            nombre = secure_filename(archivo.filename)
            ruta   = os.path.join(app.config["UPLOAD_FOLDER"], nombre)
            archivo.save(ruta)

            texto = extraer_texto_ocr(ruta)
            datos = extraer_datos(texto)
            guardar_en_excel(datos["fecha"], datos["codigo_aprobado"], datos["total"])
            resultados.append({
                "archivo":  nombre,
                **datos,
                "texto_ocr": texto[:500] + ("..." if len(texto) > 500 else ""),
            })
        else:
            flash(f"⚠️ Formato no permitido: {archivo.filename}", "warning")

    return render_template("resultado.html", resultados=resultados)


@app.route("/descargar")
def descargar():
    crear_excel_si_no_existe()
    return send_file(EXCEL_FILE, as_attachment=True, download_name="facturas.xlsx")


@app.route("/limpiar", methods=["POST"])
def limpiar():
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    flash("✅ Registro Excel limpiado correctamente.", "success")
    return redirect(url_for("index"))


def obtener_registros():
    if not os.path.exists(EXCEL_FILE):
        return []
    wb    = load_workbook(EXCEL_FILE, data_only=True)
    ws    = wb["Facturas"]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    return [{"fecha": f[0], "codigo": f[1], "total": f[2]} for f in filas if any(f)]


if __name__ == "__main__":
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    crear_excel_si_no_existe()
    app.run(debug=True, port=5000)